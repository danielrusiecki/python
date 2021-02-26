# pip install openpyxl
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

# ****************************** DANE DO UZUPELNIENIA! **********************************************************
nazwa_pliku_oryginalnego = "COLLIERS - 15 marca (poprawiony)_ORI"
nazwa_pliku_WYNIKOWEGO = "testowy_ZROBIONY_17"



# ****************************** DANE DO SPRAWDZENIA **********************************************************
# ****************************** NETTO *******************************
nr_kolumny_NETTO = 7
litera_kolumny_NETTO = "G"
# ****************************** BRUTTO ******************************
nr_kolumny_BRUTTO = 10
litera_kolumny_BRUTTO = "J"
# ****************************** NAZWISKO PASAZERA *******************
nr_kolumny_NAZWISKO_PASAZERA = 11
litera_kolumny_NAZWISKO_PASAZERA = "K"
# ****************************** CEL PODROZY *************************
nr_kolumny_CEL_PODROZY = 16
litera_kolumny_CEL_PODROZY = "P"
# ****************************** DEPARTAMENT *************************
nr_kolumny_DEPARTAMENT = 17
litera_kolumny_DEPARTAMENT = "Q"

# ****************************** NAZWISKA PASAZEROW ******************
lista_pasazer_25_75 = ("KUCHARSKI", "WLODARCZYK", "KOSCIELNIAK", "GALICKA", "BEDEKIER", "PIEKARSKI", "PRYSZCZ")
lista_pasazer_15_85 = ("MACHUS", "SZWAJA")
lista_pasazer_50_50 = ("NALEZYTY", "BLABLABLA")








# \\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
# \\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\
# \\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\

workbook = load_workbook(filename="C:\\Users\\rusiecki\\Desktop\\" + nazwa_pliku_oryginalnego + ".xlsx")
sheet = workbook.active

# licznik do przesuwania się po wierszach
rowind = 1

# sumy netto i brutto w celu finalnej weryfikacji i do pokazania na koniec programu
suma_netto_ORI = 0
suma_netto_po_podziale = 0
suma_brutto_ORI = 0
suma_brutto_po_podziale = 0

# dodanie dodatkowej kolumny na koncu tabeli i nazwanie ją CEL - FINAL
sheet.insert_cols(idx=len(sheet[1])+1)
sheet.cell(row=1, column=len(sheet[1])+1).value = "CEL - FINAL"
szerokosc_tabeli = len(sheet[1])

# uzupelnienie wlasnie dodanej kolumny polami klient/bez klienta/szkolenia w zaleznosci od tego co
# wystepuje w polu kolumny Cel podrozy
for r in range(2,len(sheet["A"])+1):
    if "KLIENT" in sheet.cell(row=r, column=nr_kolumny_CEL_PODROZY).value:
        sheet.cell(row=r, column=szerokosc_tabeli).value = "klient"
    elif "COLLIERS" in sheet.cell(row=r, column=nr_kolumny_CEL_PODROZY).value:
        sheet.cell(row=r, column=szerokosc_tabeli).value = "bez klienta"
    elif "KONFERENCJA" in sheet.cell(row=r, column=nr_kolumny_CEL_PODROZY).value:
        sheet.cell(row=r, column=szerokosc_tabeli).value = "szkolenia"
    elif "SZKOLENIE" in sheet.cell(row=r, column=nr_kolumny_CEL_PODROZY).value:
        sheet.cell(row=r, column=szerokosc_tabeli).value = "szkolenia"
    elif "" in sheet.cell(row=r, column=nr_kolumny_CEL_PODROZY).value:
        sheet.cell(row=r, column=szerokosc_tabeli).value = "bez klienta"

for row in sheet.rows:
    # sumowanie oryginalnych wartosci NETTO i BRUTTO (przed zmianami, zeby pokazac dla porownania na koncu)
    # po to ten IF aby ominac pierwszy wiersz z naglowkami
    if type(sheet.cell(row=rowind, column=nr_kolumny_NETTO).value) != str:
        suma_netto_ORI += sheet.cell(row=rowind, column=nr_kolumny_NETTO).value
        suma_brutto_ORI += sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value

    # wlasciwy FOR dla konkretnych list pasazerow
    for pasazer in lista_pasazer_25_75:
        if pasazer in sheet.cell(row=rowind, column=nr_kolumny_NAZWISKO_PASAZERA).value:
            sheet.insert_rows(idx=rowind)

            # kopiowanie kazdej komorki z wiersza nizszego do aktualnego (wyzszego)
            for i in range(1,len(sheet[rowind])+1):
                sheet.cell(row=rowind, column=i).value = sheet.cell(row=rowind + 1, column=i).value

            # w gornym wierszu mnozy NETTO oraz BRUTTO razy 0.25 a w dolnym razy 0.75
            sheet.cell(row=rowind, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind, column=nr_kolumny_NETTO).value * 0.25
            sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value * 0.25
            sheet.cell(row=rowind+1, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind+1, column=nr_kolumny_NETTO).value * 0.75
            sheet.cell(row=rowind+1, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind+1, column=nr_kolumny_BRUTTO).value * 0.75

            # sumowanie wartosci NETTO i BRUTTO juz po podziale (do pokazania na koncu programu)
            # suma_netto_po_podziale += sheet.cell(row=rowind, column=nr_kolumny_NETTO).value
            # suma_netto_po_podziale += sheet.cell(row=rowind+1, column=nr_kolumny_NETTO).value
            # suma_brutto_po_podziale += sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value
            # suma_brutto_po_podziale += sheet.cell(row=rowind+1, column=nr_kolumny_BRUTTO).value


            # zamiana OFFICE TENANT na ADMINISTRATION jesli w gornym wierszu jest OFFICE TENANT
            if "OFFICE TENANT" in sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind, column=nr_kolumny_DEPARTAMENT).value.replace("OFFICE TENANT", "ADMINISTRATION")

                # zamiana ADMINISTRATION na OFFICE TENANT jesli w dolnym wierszu jest ADMINISTRATION
            if "ADMINISTRATION" in sheet.cell(row=rowind+1, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind+1, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind+1, column=nr_kolumny_DEPARTAMENT).value.replace("ADMINISTRATION", "OFFICE TENANT")

            # dodatkowe zwiekszenie licznika bo wstawilismy wiersz wiec trzeba go ominac
            rowind += 1

    for pasazer in lista_pasazer_15_85:
        if pasazer in sheet.cell(row=rowind, column=nr_kolumny_NAZWISKO_PASAZERA).value:
            sheet.insert_rows(idx=rowind)

            # kopiowanie kazdej komorki z wiersza nizszego do aktualnego (wyzszego)
            for i in range(1, len(sheet[rowind]) + 1):
                sheet.cell(row=rowind, column=i).value = sheet.cell(row=rowind + 1, column=i).value

            # w gornym wierszu mnozy NETTO oraz BRUTTO razy 0.25 a w dolnym razy 0.75
            sheet.cell(row=rowind, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind, column=nr_kolumny_NETTO).value * 0.15
            sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value * 0.15
            sheet.cell(row=rowind + 1, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind + 1, column=nr_kolumny_NETTO).value * 0.85
            sheet.cell(row=rowind + 1, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind + 1, column=nr_kolumny_BRUTTO).value * 0.85

            # zamiana OFFICE TENANT na ADMINISTRATION jesli w gornym wierszu jest OFFICE TENANT
            if "OFFICE TENANT" in sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind, column=nr_kolumny_DEPARTAMENT).value.replace("OFFICE TENANT", "ADMINISTRATION")

                # zamiana ADMINISTRATION na OFFICE TENANT jesli w dolnym wierszu jest ADMINISTRATION
            if "ADMINISTRATION" in sheet.cell(row=rowind + 1, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind + 1, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind + 1, column=nr_kolumny_DEPARTAMENT).value.replace("ADMINISTRATION", "OFFICE TENANT")

            rowind += 1

    for pasazer in lista_pasazer_50_50:
        if pasazer in sheet.cell(row=rowind, column=nr_kolumny_NAZWISKO_PASAZERA).value:
            sheet.insert_rows(idx=rowind)

            # kopiowanie kazdej komorki z wiersza nizszego do aktualnego (wyzszego)
            for i in range(1,len(sheet[rowind])+1):
                sheet.cell(row=rowind, column=i).value = sheet.cell(row=rowind + 1, column=i).value

            # w gornym wierszu mnozy NETTO oraz BRUTTO razy 0.25 a w dolnym razy 0.75
            sheet.cell(row=rowind, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind, column=nr_kolumny_NETTO).value * 0.5
            sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value * 0.5
            sheet.cell(row=rowind+1, column=nr_kolumny_NETTO).value = sheet.cell(row=rowind+1, column=nr_kolumny_NETTO).value * 0.5
            sheet.cell(row=rowind+1, column=nr_kolumny_BRUTTO).value = sheet.cell(row=rowind+1, column=nr_kolumny_BRUTTO).value * 0.5

            # zamiana OFFICE TENANT na ADMINISTRATION jesli w gornym wierszu jest OFFICE TENANT
            if "OFFICE TENANT" in sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind, column=nr_kolumny_DEPARTAMENT).value.replace("OFFICE TENANT", "ADMINISTRATION")

                # zamiana ADMINISTRATION na OFFICE TENANT jesli w dolnym wierszu jest ADMINISTRATION
            if "ADMINISTRATION" in sheet.cell(row=rowind+1, column=nr_kolumny_DEPARTAMENT).value:
                sheet.cell(row=rowind+1, column=nr_kolumny_DEPARTAMENT).value = sheet.cell(
                    row=rowind+1, column=nr_kolumny_DEPARTAMENT).value.replace("ADMINISTRATION", "OFFICE TENANT")

            rowind += 1


    # przejście do kolejnego wiersza
    rowind += 1

    # if type(sheet.cell(row=rowind, column=nr_kolumny_NETTO).value) != str:
    #     suma_netto_po_podziale += sheet.cell(row=rowind, column=nr_kolumny_NETTO).value
    #     suma_brutto_po_podziale += sheet.cell(row=rowind, column=nr_kolumny_BRUTTO).value
# *******************************************************************************************************************
# suma wartosci NETTO i BRUTTO po podziale (do sprawdzenia i pokazania na koncu programu)
for i in range(2,len(sheet["A"])+1):
    if type(sheet.cell(row=i, column=nr_kolumny_NETTO).value) != str:
        suma_netto_po_podziale += sheet.cell(row=i, column=nr_kolumny_NETTO).value
        suma_brutto_po_podziale += sheet.cell(row=i, column=nr_kolumny_BRUTTO).value
# *******************************************************************************************************************
pogrubiony = Font(bold=True)
liczba_wierszy = len(sheet["A"])

# SUMA NETTO NA KONCU KOLUMNY NETTO
formula_sum_netto = "=SUM(" + litera_kolumny_NETTO + "2:" + litera_kolumny_NETTO + str(liczba_wierszy) + ")"
# print(formula_sum_netto)
sheet.cell(row=liczba_wierszy+1, column=nr_kolumny_NETTO).value = formula_sum_netto
sheet.cell(row=liczba_wierszy+1, column=nr_kolumny_NETTO).font = pogrubiony

# SUMA BRUTTO NA KONCU KOLUMNY BRUTTO
formula_sum_brutto = "=SUM(" + litera_kolumny_BRUTTO + "2:" + litera_kolumny_BRUTTO + str(liczba_wierszy) + ")"
# print(formula_sum_brutto)
sheet.cell(row=liczba_wierszy+1, column=nr_kolumny_BRUTTO).value = formula_sum_brutto
sheet.cell(row=liczba_wierszy+1, column=nr_kolumny_BRUTTO).font = Font(bold=True)

# sumy PRZED podzialem
# print(round(suma_netto_ORI, 2))
# print(round(suma_brutto_ORI, 2))

# sumy PO podziale
print("Suma NETTO po podziale wynosi:")
print(round(suma_netto_po_podziale, 2))
print("\n")
print("Suma BRUTTO po podziale wynosi:")
print(round(suma_brutto_po_podziale, 2))
print("\n")
print("********************************************************************************")

# workbook.save(filename="C:\\Users\\rusiecki\\Desktop\\" + nazwa_pliku_WYNIKOWEGO + ".xlsx")

os.system("pause")